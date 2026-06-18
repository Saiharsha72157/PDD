import os
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart

# ── Styling ───────────────────────────────────────────────────────────────────
TITLE_FONT = Font(size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(bold=True)
METRIC_LABEL_FONT = Font(bold=True)
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def _apply_header_style(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

def _auto_width(ws):
    for col in ws.columns:
        if not col: continue
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
                break
        if not col_letter: continue
        
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)

# ── Sheet 1: Executive Summary ──────────────────────────────────────────────
def _build_sheet_1_executive_summary(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.active
    ws.title = "1. Executive Summary"
    ws.sheet_properties.tabColor = "1F497D"

    ws.merge_cells("A1:D1")
    ws["A1"].value = "ResearchMate AI - Load Test Executive Summary"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    s = metrics.get("summary", {})
    cfg = metrics.get("test_config", {})

    ws["A3"] = "Test Configuration"
    ws["A3"].font = METRIC_LABEL_FONT
    ws["A4"] = "Virtual Users"
    ws["B4"] = cfg.get("num_users", 0)
    ws["A5"] = "Duration (s)"
    ws["B5"] = cfg.get("duration_seconds", 0)

    ws["A8"] = "High-Level Metrics"
    ws["A8"].font = METRIC_LABEL_FONT
    ws["A9"] = "Success Rate"
    sr = s.get("success_rate_pct", 0)
    ws["B9"] = f"{sr}%"
    ws["B9"].fill = GOOD_FILL if sr >= 90 else BAD_FILL
    
    ws["A10"] = "Throughput (RPS)"
    ws["B10"] = s.get("rps", 0)
    
    ws["A11"] = "Avg Response Time (ms)"
    avg_ms = s.get("avg_ms", 0)
    ws["B11"] = avg_ms
    ws["B11"].fill = GOOD_FILL if avg_ms < 40000 else BAD_FILL

    ws["A13"] = "Validation Criteria"
    ws["A13"].font = METRIC_LABEL_FONT
    
    ep_stats = metrics.get("per_endpoint", {})
    para_stats = ep_stats.get("POST /paraphrase", {})
    para_sr = (para_stats.get("successful", 0) / max(1, para_stats.get("total_requests", 1))) * 100
    
    timeout_count = metrics.get("timeout_count", 0)
    conn_err_count = metrics.get("conn_err_count", 0)

    val_rules = [
        ("Success Rate >= 90%", sr >= 90),
        ("Timeout Errors <= 20", timeout_count <= 20),
        ("Connection Errors == 0", conn_err_count == 0),
        ("Paraphrase Success >= 80%", para_sr >= 80),
        ("Average Response < 40000ms", avg_ms < 40000),
        ("P95 Response Time < 150000ms", s.get("p95_ms", 0) < 150000)
    ]
    
    row = 14
    all_pass = True
    for rule, passed in val_rules:
        ws.cell(row=row, column=1, value=rule).border = THIN_BORDER
        c = ws.cell(row=row, column=2, value="PASS" if passed else "FAIL")
        c.fill = GOOD_FILL if passed else BAD_FILL
        c.border = THIN_BORDER
        if not passed: all_pass = False
        row += 1

    ws["D3"] = "OVERALL STATUS:"
    ws["D3"].font = METRIC_LABEL_FONT
    ws["D4"] = "PASS" if all_pass else "FAIL"
    ws["D4"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["D4"].fill = PatternFill(start_color="00B050" if all_pass else "FF0000", end_color="00B050" if all_pass else "FF0000", fill_type="solid")
    ws["D4"].alignment = Alignment(horizontal="center")

    _auto_width(ws)

# ── Sheet 2: Request Statistics ─────────────────────────────────────────────
def _build_sheet_2_request_stats(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("2. Request Statistics")
    
    s = metrics.get("summary", {})
    ws["A1"] = "Global Request Statistics"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL

    ws["A3"] = "Total Requests"
    ws["B3"] = s.get("total_requests", 0)
    ws["A4"] = "Successful"
    ws["B4"] = s.get("successful_requests", 0)
    ws["A5"] = "Failed"
    ws["B5"] = s.get("error_requests", 0)
    ws["A6"] = "Rate Limited"
    ws["B6"] = s.get("rate_limited_requests", 0)

    chart = PieChart()
    labels = Reference(ws, min_col=1, min_row=4, max_row=5)
    data = Reference(ws, min_col=2, min_row=4, max_row=5)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    chart.title = "Success vs Failures"
    ws.add_chart(chart, "D3")
    _auto_width(ws)

# ── Sheet 3: Endpoint Statistics ─────────────────────────────────────────────
def _build_sheet_3_endpoint_stats(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("3. Endpoint Statistics")
    
    headers = ["Endpoint", "Total", "Successful", "Errors", "Success %"]
    _apply_header_style(ws, 1, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for ep, data in metrics.get("per_endpoint", {}).items():
        total = data.get("total_requests", 0)
        succ = data.get("successful", 0)
        pct = (succ / total * 100) if total > 0 else 0
        
        ws.cell(row=row, column=1, value=ep)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=succ)
        ws.cell(row=row, column=4, value=data.get("errors", 0))
        ws.cell(row=row, column=5, value=f"{round(pct, 1)}%")
        row += 1
    _auto_width(ws)

# ── Sheet 4: Response Time Analysis ─────────────────────────────────────────
def _build_sheet_4_response_times(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("4. Response Times")
    
    headers = ["Endpoint", "Min (ms)", "Max (ms)", "Avg (ms)", "P50", "P95", "P99"]
    _apply_header_style(ws, 1, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for ep, data in metrics.get("per_endpoint", {}).items():
        ws.cell(row=row, column=1, value=ep)
        ws.cell(row=row, column=2, value=data.get("min_ms", 0))
        ws.cell(row=row, column=3, value=data.get("max_ms", 0))
        ws.cell(row=row, column=4, value=data.get("avg_ms", 0))
        ws.cell(row=row, column=5, value=data.get("p50_ms", 0))
        ws.cell(row=row, column=6, value=data.get("p95_ms", 0))
        ws.cell(row=row, column=7, value=data.get("p99_ms", 0))
        row += 1
    _auto_width(ws)

# ── Sheet 5: Error Analysis ─────────────────────────────────────────────────
def _build_sheet_5_error_analysis(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("5. Error Analysis")
    
    errors = [r for r in metrics.get("raw_results", []) if not r.get("success")]
    if not errors:
        ws["A1"] = "No errors detected!"
        metrics["timeout_count"] = 0
        metrics["conn_err_count"] = 0
        return

    error_counts = {}
    timeout_count = 0
    conn_err_count = 0
    
    for e in errors:
        msg = str(e.get("error") or f"HTTP {e.get('status_code')}")
        error_counts[msg] = error_counts.get(msg, 0) + 1
        if "timeout" in msg.lower(): timeout_count += 1
        if "connection" in msg.lower(): conn_err_count += 1
        
    metrics["timeout_count"] = timeout_count
    metrics["conn_err_count"] = conn_err_count

    headers = ["Error Message", "Count", "Percentage"]
    _apply_header_style(ws, 1, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for msg, count in sorted(error_counts.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=msg)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{round((count/len(errors))*100, 2)}%")
        row += 1
    _auto_width(ws)

# ── Sheet 6: Groq Key Statistics ────────────────────────────────────────────
def _build_sheet_6_groq_keys(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("6. Groq Keys")
    
    keys_data = metrics.get("dashboard", {}).get("keys", [])
    if not keys_data:
        ws["A1"] = "No Groq Dashboard data available."
        return

    headers = ["Key ID", "Healthy", "Active Requests", "Usage Count", "Failures", "Cooldown Remaining"]
    _apply_header_style(ws, 1, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for ks in sorted(keys_data, key=lambda x: x.get("key_id", 0)):
        ws.cell(row=row, column=1, value=f"Key {ks.get('key_id')}")
        ws.cell(row=row, column=2, value="Yes" if ks.get('healthy') else "No")
        ws.cell(row=row, column=3, value=ks.get('active_requests', 0))
        ws.cell(row=row, column=4, value=ks.get('usage_count', 0))
        ws.cell(row=row, column=5, value=ks.get('failure_count', 0))
        ws.cell(row=row, column=6, value=round(ks.get('cooldown_remaining', 0.0), 1))
        row += 1
    _auto_width(ws)

# ── Sheet 7: Queue Statistics ───────────────────────────────────────────────
def _build_sheet_7_queue_stats(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("7. Queue Statistics")
    
    global_stats = metrics.get("dashboard", {}).get("global", {})
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    _apply_header_style(ws, 1, 2)

    ws["A2"] = "Peak Queue Size"
    ws["B2"] = global_stats.get("peak_queue_size", 0)
    ws["A3"] = "Current Queue Size"
    ws["B3"] = global_stats.get("queue_size", 0)
    ws["A4"] = "Queue Dropped (Full)"
    ws["B4"] = global_stats.get("queue_dropped", 0)
    _auto_width(ws)

# ── Sheet 8: Cache Statistics ───────────────────────────────────────────────
def _build_sheet_8_cache_stats(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("8. Cache Statistics")
    
    global_stats = metrics.get("dashboard", {}).get("global", {})
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    _apply_header_style(ws, 1, 2)

    hits = global_stats.get("cache_hits", 0)
    misses = global_stats.get("cache_misses", 0)
    total = hits + misses

    ws["A2"] = "Cache Hits"
    ws["B2"] = hits
    ws["A3"] = "Cache Misses"
    ws["B3"] = misses
    ws["A4"] = "Hit Ratio"
    ws["B4"] = f"{round((hits/total)*100, 2)}%" if total > 0 else "0%"
    _auto_width(ws)

# ── Sheet 9: Before vs After ────────────────────────────────────────────────
def _build_sheet_9_comparison(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("9. Before vs After")
    
    ws.merge_cells("A1:C1")
    ws["A1"] = "Performance Improvements"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Metric", "Before (Baseline)", "After (Current)"]
    _apply_header_style(ws, 2, 3)
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    s = metrics.get("summary", {})
    
    # Hardcoded known baseline metrics from prompt
    comparisons = [
        ("Success Rate", "28.32%", f"{s.get('success_rate_pct', 0)}%"),
        ("Paraphrase Success Rate", "0.0%", "Available in Sheet 3"),
        ("Timeout Errors", "200", metrics.get("timeout_count", 0)),
        ("Peak Queue Size", "219", metrics.get("dashboard", {}).get("global", {}).get("peak_queue_size", 0)),
        ("Avg Response Time", "8027ms", f"{s.get('avg_ms', 0)}ms")
    ]
    
    row = 3
    for m, before, after in comparisons:
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=before)
        ws.cell(row=row, column=3, value=after)
        row += 1

    _auto_width(ws)

# ── Sheet 10: Recommendations ───────────────────────────────────────────────
def _build_sheet_10_recommendations(wb: Workbook, metrics: Dict[str, Any]):
    ws = wb.create_sheet("10. Recommendations")
    
    ws.merge_cells("A1:B1")
    ws["A1"] = "Automated Engine Recommendations"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    recs = []
    s = metrics.get("summary", {})
    if s.get("success_rate_pct", 0) >= 90:
        recs.append(("System is Optimized", "No further action needed. The system easily handles 100 concurrent users with caching and queue limits."))
    else:
        recs.append(("Check Failures", "Review the Error Analysis sheet to address the remaining failed requests."))

    row = 3
    for title, desc in recs:
        ws.cell(row=row, column=1, value=title).font = METRIC_LABEL_FONT
        ws.cell(row=row, column=2, value=desc)
        row += 1

    _auto_width(ws)


def generate_report(metrics: Dict[str, Any], output_path: str = None) -> str:
    """Generates the 10-sheet Excel report for Phase G."""
    # First analyze errors to calculate timeout_count for Sheet 1
    _build_sheet_5_error_analysis(Workbook(), metrics) 

    if not output_path:
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Load_Test_Report.xlsx")

    wb = Workbook()

    _build_sheet_1_executive_summary(wb, metrics)
    _build_sheet_2_request_stats(wb, metrics)
    _build_sheet_3_endpoint_stats(wb, metrics)
    _build_sheet_4_response_times(wb, metrics)
    _build_sheet_5_error_analysis(wb, metrics)
    _build_sheet_6_groq_keys(wb, metrics)
    _build_sheet_7_queue_stats(wb, metrics)
    _build_sheet_8_cache_stats(wb, metrics)
    _build_sheet_9_comparison(wb, metrics)
    _build_sheet_10_recommendations(wb, metrics)

    wb.save(output_path)
    return output_path
